#!/usr/bin/env python3
"""Agente semanal Apollo → BBDD → HubSpot — Estrategia Mobiliario Urbano.

Cada semana:
  1. Elige un lote de municipios de data/contactos.csv, priorizando los que
     aún no tienen ningún contacto nominal (nombre + cargo + email).
  2. Busca en Apollo la organización «Ayuntamiento de <municipio>» y las
     personas con cargos de decisión (agent/config.json → cargos_objetivo).
  3. Enriquece cada persona para obtener su email verificado.
  4. Añade los contactos nuevos a data/contactos.csv (origen=apollo) y
     regenera data/ayuntamientos-costeros.xlsx.
  5. Da de alta cada contacto nuevo en HubSpot (portal 26243090) con las
     propiedades del grupo «Mobiliario Urbano», para que el workflow del
     embudo continúe desde ahí.

Uso:
  export APOLLO_API_KEY=...
  export HUBSPOT_PRIVATE_APP_TOKEN=pat-eu1-...   # opcional: sin él no sincroniza
  python3 agent/apollo_agent.py [--dry-run] [--lote N] [--solo-export-xlsx]

--dry-run busca y muestra lo que haría, pero no escribe CSV/xlsx ni toca HubSpot.
"""

import argparse
import csv
import datetime
import json
import os
import sys
import time
import unicodedata
import urllib.error
import urllib.request

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(REPO, "data", "contactos.csv")
XLSX_PATH = os.path.join(REPO, "data", "ayuntamientos-costeros.xlsx")
STATE_PATH = os.path.join(REPO, "agent", "state.json")
CONFIG_PATH = os.path.join(REPO, "agent", "config.json")

APOLLO_API = "https://api.apollo.io/api/v1"
HUBSPOT_API = "https://api.hubapi.com"

FIELDNAMES = [
    "municipio", "provincia", "comunidad_autonoma", "nombre", "apellidos",
    "cargo", "email", "tipo_email", "fuente", "origen", "fecha_alta",
    "hubspot_sync",
]


# ---------------------------------------------------------------- utilidades

def norm(s):
    """minúsculas y sin tildes, para comparar municipios y emails."""
    s = (s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def http_json(method, url, headers=None, body=None, retries=3):
    data = json.dumps(body).encode() if body is not None else None
    hdrs = {"Content-Type": "application/json", "Accept": "application/json"}
    hdrs.update(headers or {})
    for intento in range(retries):
        req = urllib.request.Request(url, data=data, headers=hdrs, method=method)
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return resp.status, json.loads(resp.read().decode() or "{}")
        except urllib.error.HTTPError as e:
            cuerpo = e.read().decode(errors="replace")
            if e.code == 429 or e.code >= 500:  # rate limit / caída transitoria
                espera = 2 ** (intento + 1)
                print(f"    HTTP {e.code}, reintento en {espera}s…")
                time.sleep(espera)
                continue
            return e.code, {"error": cuerpo}
        except urllib.error.URLError as e:
            if intento < retries - 1:
                time.sleep(2 ** (intento + 1))
                continue
            return 0, {"error": str(e)}
    return 0, {"error": "reintentos agotados"}


def cargar_csv():
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def guardar_csv(filas):
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(filas)


# ------------------------------------------------------------------- Apollo

def apollo_headers():
    return {"X-Api-Key": os.environ["APOLLO_API_KEY"],
            "Cache-Control": "no-cache"}


def buscar_organizacion(municipio):
    """Devuelve el id de organización Apollo del ayuntamiento, o None."""
    candidatos = [f"Ayuntamiento de {municipio}", f"Ajuntament de {municipio}",
                  f"Concello de {municipio}", f"{municipio} City Council", municipio]
    for nombre in candidatos:
        status, data = http_json(
            "POST", f"{APOLLO_API}/mixed_companies/search",
            headers=apollo_headers(),
            body={"q_organization_name": nombre, "per_page": 5,
                  "organization_locations": ["Spain"]})
        if status != 200:
            print(f"    Apollo companies/search → HTTP {status}: {data.get('error', '')[:200]}")
            return None
        objetivo = norm(municipio)
        for org in (data.get("organizations") or []) + (data.get("accounts") or []):
            n = norm(org.get("name"))
            if objetivo in n and any(p in n for p in ("ayuntamiento", "ajuntament", "ajuntamento", "city council", "concello", objetivo)):
                return org.get("organization_id") or org.get("id")
    return None


def buscar_personas(org_id, cargos, max_personas):
    status, data = http_json(
        "POST", f"{APOLLO_API}/mixed_people/search",
        headers=apollo_headers(),
        body={"organization_ids": [org_id], "person_titles": cargos,
              "per_page": max(max_personas * 2, 10), "page": 1})
    if status != 200:
        print(f"    Apollo people/search → HTTP {status}: {data.get('error', '')[:200]}")
        return []
    return (data.get("people") or [])[: max_personas * 2]


def enriquecer_persona(persona):
    """people/match consume 1 crédito y devuelve el email desbloqueado."""
    status, data = http_json(
        "POST", f"{APOLLO_API}/people/match",
        headers=apollo_headers(),
        body={"id": persona.get("id"),
              "reveal_personal_emails": False})
    if status != 200:
        return None
    p = data.get("person") or {}
    email = p.get("email")
    if not email or email.endswith("@domain.com") or "not_unlocked" in email:
        return None
    return {"nombre": p.get("first_name") or persona.get("first_name") or "",
            "apellidos": p.get("last_name") or persona.get("last_name") or "",
            "cargo": p.get("title") or persona.get("title") or "",
            "email": email,
            "linkedin": p.get("linkedin_url") or ""}


# ------------------------------------------------------------------ HubSpot

def hubspot_upsert(contacto, cfg):
    """Alta/actualización idempotente del contacto por email en HubSpot."""
    token = os.environ.get("HUBSPOT_PRIVATE_APP_TOKEN")
    if not token:
        return "sin_token"
    props = {
        "email": contacto["email"],
        "firstname": contacto["nombre"],
        "lastname": contacto["apellidos"],
        "jobtitle": contacto["cargo"],
        "company": f"Ayuntamiento de {contacto['municipio']}",
        "tipo_entidad": cfg["hubspot"]["tipo_entidad"],
        "municipio": contacto["municipio"],
        "canal_origen": cfg["hubspot"]["canal_origen"],
    }
    headers = {"Authorization": f"Bearer {token}"}
    status, data = http_json(
        "POST", f"{HUBSPOT_API}/crm/v3/objects/contacts",
        headers=headers, body={"properties": props})
    if status == 201:
        return "creado"
    if status == 409:  # ya existe → actualizar por email
        status, data = http_json(
            "PATCH",
            f"{HUBSPOT_API}/crm/v3/objects/contacts/{contacto['email']}?idProperty=email",
            headers=headers, body={"properties": props})
        return "actualizado" if status == 200 else f"error_{status}"
    print(f"    HubSpot → HTTP {status}: {str(data)[:200]}")
    return f"error_{status}"


# --------------------------------------------------------------------- xlsx

def exportar_xlsx(filas):
    """Regenera el Excel (hoja de datos + Resumen) desde el CSV canónico."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ayuntamientos costeros"
    cab = ["Municipio", "Provincia", "Comunidad Autónoma", "Nombre", "Apellidos",
           "Cargo", "Email", "Tipo de email", "Fuente (URL)", "Origen", "Fecha alta"]
    ws.append(cab)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True)
    for f in filas:
        ws.append([f["municipio"], f["provincia"], f["comunidad_autonoma"],
                   f["nombre"], f["apellidos"], f["cargo"], f["email"],
                   f["tipo_email"], f["fuente"], f["origen"], f["fecha_alta"]])
    anchos = [24, 14, 20, 14, 18, 30, 34, 22, 40, 10, 12]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = a
    for fila in ws.iter_rows(min_row=2):
        for c in fila:
            c.font = Font(name="Arial")

    municipios = {norm(f["municipio"]) for f in filas}
    con_email = {norm(f["municipio"]) for f in filas if "@" in f["email"]}
    res = wb.create_sheet("Resumen")
    datos = [
        ("Base de datos: Ayuntamientos costeros (Cataluña, C. Valenciana, Baleares, Canarias)", None),
        (None, None),
        ("Total de municipios", len(municipios)),
        ("Municipios con email localizado", len(con_email)),
        ("Municipios sin email público localizado", len(municipios) - len(con_email)),
        ("Contactos totales", sum(1 for f in filas if "@" in f["email"])),
        ("Contactos nominales (Apollo)", sum(1 for f in filas if f["origen"] == "apollo")),
        (None, None),
        ("Última actualización del agente", datetime.date.today().isoformat()),
    ]
    for r in datos:
        res.append(r)
    res["A1"].font = Font(name="Arial", bold=True)
    for fila in res.iter_rows():
        for c in fila:
            if not c.font.bold:
                c.font = Font(name="Arial")
    res.column_dimensions["A"].width = 60
    res.column_dimensions["B"].width = 12
    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------- run

def elegir_lote(filas, tamano):
    """Municipios sin contacto nominal primero; rotación estable por estado."""
    con_nominal = {norm(f["municipio"]) for f in filas if f["origen"] == "apollo" and f["email"]}
    municipios = []  # únicos, en orden del CSV
    vistos = set()
    for f in filas:
        m = norm(f["municipio"])
        if m not in vistos:
            vistos.add(m)
            municipios.append((f["municipio"], f["provincia"], f["comunidad_autonoma"]))

    estado = {}
    if os.path.exists(STATE_PATH):
        with open(STATE_PATH, encoding="utf-8") as f:
            estado = json.load(f)
    procesados = set(estado.get("procesados", []))

    con_email = {norm(f["municipio"]) for f in filas if "@" in f["email"]}
    pendientes = [m for m in municipios
                  if norm(m[0]) not in con_nominal and norm(m[0]) not in procesados]
    if not pendientes:  # vuelta completa: reiniciar la rotación
        procesados = set()
        pendientes = [m for m in municipios if norm(m[0]) not in con_nominal]
    # primero los municipios sin ningún email; dentro de cada grupo, orden del CSV
    pendientes.sort(key=lambda m: norm(m[0]) in con_email)
    return pendientes[:tamano], procesados, estado


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--lote", type=int, default=None,
                    help="municipios a procesar en esta ejecución")
    ap.add_argument("--solo-export-xlsx", action="store_true",
                    help="regenera el Excel desde el CSV y termina")
    args = ap.parse_args()

    with open(CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    filas = cargar_csv()

    if args.solo_export_xlsx:
        exportar_xlsx(filas)
        print(f"Excel regenerado: {XLSX_PATH}")
        return

    if not os.environ.get("APOLLO_API_KEY"):
        sys.exit("Falta APOLLO_API_KEY en el entorno")

    lote, procesados, estado = elegir_lote(
        filas, args.lote or cfg["municipios_por_ejecucion"])
    print(f"Lote de esta semana: {len(lote)} municipios → "
          + ", ".join(m[0] for m in lote))

    emails_existentes = {norm(f["email"]) for f in filas if "@" in f["email"]}
    hoy = datetime.date.today().isoformat()
    nuevos = []

    for municipio, provincia, ca in lote:
        print(f"\n· {municipio} ({provincia})")
        procesados.add(norm(municipio))
        org_id = buscar_organizacion(municipio)
        if not org_id:
            print("    sin organización en Apollo")
            continue
        personas = buscar_personas(org_id, cfg["cargos_objetivo"],
                                   cfg["max_contactos_por_municipio"])
        anadidos = 0
        for p in personas:
            if anadidos >= cfg["max_contactos_por_municipio"]:
                break
            enr = enriquecer_persona(p)
            if not enr or norm(enr["email"]) in emails_existentes:
                continue
            emails_existentes.add(norm(enr["email"]))
            fila = {"municipio": municipio, "provincia": provincia,
                    "comunidad_autonoma": ca, "nombre": enr["nombre"],
                    "apellidos": enr["apellidos"], "cargo": enr["cargo"],
                    "email": enr["email"], "tipo_email": "Nominal (Apollo)",
                    "fuente": enr["linkedin"] or "apollo.io",
                    "origen": "apollo", "fecha_alta": hoy, "hubspot_sync": ""}
            nuevos.append(fila)
            anadidos += 1
            print(f"    + {enr['nombre']} {enr['apellidos']} — {enr['cargo']} <{enr['email']}>")
        if not anadidos:
            print("    sin contactos nuevos con email")

    print(f"\nContactos nuevos: {len(nuevos)}")
    if args.dry_run:
        print("(dry-run: no se escribe nada)")
        return

    for fila in nuevos:
        fila["hubspot_sync"] = hubspot_upsert(fila, cfg)
    filas.extend(nuevos)
    guardar_csv(filas)
    exportar_xlsx(filas)

    estado["procesados"] = sorted(procesados)
    estado["ultima_ejecucion"] = hoy
    estado["nuevos_ultima_ejecucion"] = len(nuevos)
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)

    sincronizados = sum(1 for f in nuevos if f["hubspot_sync"] in ("creado", "actualizado"))
    print(f"CSV y Excel actualizados · HubSpot: {sincronizados}/{len(nuevos)} sincronizados")


if __name__ == "__main__":
    main()
